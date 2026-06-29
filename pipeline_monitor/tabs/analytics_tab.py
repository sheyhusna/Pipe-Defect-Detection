"""
tabs/analytics_tab.py
──────────────────────
Analytics dashboard with charts, structured Excel export, and scrollable gallery.
"""

import os
import json
import base64
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from charts import build_confidence_chart, build_fps_chart, build_class_donut
from components import render_section_label


def render_analytics_tab():
    """Render the Analytics tab content."""

    # ── BULLETPROOF LOADING: Recover data if Streamlit wiped the session ──
    if st.session_state.total_frames == 0 and os.path.exists("results.json"):
        try:
            with open("results.json", "r") as f:
                data = json.load(f)
                st.session_state.total_frames = data.get("total_frames", 0)
                st.session_state.corrosion_fields = data.get("corrosion_frames", 0)
                st.session_state.pipe_frames = data.get("pipe_frames", 0)
        except Exception:
            pass # Ignore errors and use 0

    render_section_label("Session analytics")

    col_a, col_b, col_c = st.columns(3)

    with col_a:
        st.altair_chart(
            build_class_donut(
                st.session_state.corrosion_frames,
                st.session_state.pipe_frames,
            )
        )

    with col_b:
        fps_history = [1000 / ms for ms in st.session_state.frame_times if ms > 0]
        st.altair_chart(build_fps_chart(fps_history))

    with col_c:
        st.altair_chart(
            build_confidence_chart(list(st.session_state.conf_history))
        )

    st.markdown("---")

    # ── Structured Excel Export ─────────────────────────────────────────────
    render_section_label("Export detection log (Engineer Report)")

    if st.session_state.log_entries:
        log_df = pd.DataFrame(st.session_state.log_entries)
        report_df = log_df[["timestamp", "frame", "source", "labels", "max_conf", "fps", "is_corrosion"]].copy()
        report_df.columns = ["Timestamp", "Frame #", "Source", "Detected Objects", "max_conf", "Inference FPS", "Corrosion Alert"]

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Detection Log"

        for r_idx, row in enumerate(dataframe_to_rows(report_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center")
                if r_idx > 1 and report_df.iloc[r_idx-2]["Corrosion Alert"] == True:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output)
        excel_data = output.getvalue()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="⬇  Download Excel Report (.xlsx)",
            data=excel_data,
            file_name=f"Pipeline_Inspection_Report_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.caption("No log entries yet — run detection first.")

    # ── Filmstrip Image Gallery ────────────────────────────────────────────
    st.markdown("---")
    render_section_label("Saved Defect Frames (Scroll horizontally, click to zoom)", margin_top="1rem")

    saved_files = [f for f in st.session_state.saved_frames if os.path.exists(f)]

    if not saved_files:
        st.caption("Auto-saved corrosion frames will appear here.")
        return

    # Build gallery HTML ONLY (No modal here!)
    film_items = ""
    for filepath in saved_files:
        try:
            with open(filepath, "rb") as image_file:
                img_data = base64.b64encode(image_file.read()).decode('utf-8')

            filename = os.path.basename(filepath)
            name_part = filename.replace(".jpg", "").replace("_", " ").upper()

            # Notice the parent.document! This breaks out of the iframe to find the global modal
            film_items += f'''<div class="film-item" onclick="parent.openFS('{img_data}', '{name_part}')"><img src="data:image/jpeg;base64,{img_data}"><div class="film-label">{name_part}</div></div>
'''
        except Exception as e:
            pass

    # Container for the gallery grid
    gallery_html = f'''
<style>
.filmstrip-container {{ width: 100%; overflow-x: auto; padding: 10px 0 20px 0; }}
.filmstrip-container::-webkit-scrollbar {{ height: 8px; }}
.filmstrip-container::-webkit-scrollbar-track {{ background: #161B22; border-radius: 4px; }}
.filmstrip-container::-webkit-scrollbar-thumb {{ background: #30363D; border-radius: 4px; }}
.filmstrip {{ display: flex; flex-direction: row; gap: 15px; width: max-content; }}
.film-item {{ flex: 0 0 auto; width: 160px; cursor: pointer; border-radius: 8px; overflow: hidden; border: 2px solid #21262D; transition: all 0.2s ease; background: #0D1117; }}
.film-item:hover {{ border-color: #79C0FF; transform: translateY(-5px); box-shadow: 0 8px 15px rgba(0,0,0,0.5); }}
.film-item img {{ width: 100%; height: 100px; object-fit: cover; display: block; }}
.film-label {{ padding: 8px; font-size: 10px; color: #C9D1D9; font-family: 'JetBrains Mono', monospace; text-align: center; line-height: 1.3; white-space: pre-line; border-top: 1px solid #21262D; }}
</style>

<div class="filmstrip-container">
<div class="filmstrip">
{film_items}
</div>
</div>
'''

    # Render the gallery in an iframe
    components.html(gallery_html, height=350)

    # ── THE ZOOM MODAL (Outside the iframe, in the main document!) ──────────
    st.markdown('''
<div id="imgModal" class="modal" onclick="closeFS(event)">
      <span class="modal-close">&times;</span>
      <img class="modal-content" id="modalImg">
      <div class="modal-caption" id="modalCap"></div>
    </div>

    <script>
    function openFS(data, text) {{
        parent.document.getElementById('modalImg').src = 'data:image/jpeg;base64,' + data;
        parent.document.getElementById('modalCap').textContent = text;
        parent.document.getElementById('imgModal').style.display = 'block';
    }}
    function closeFS(e) {{
        if (e.target !== parent.document.getElementById('modalImg')) {{
            parent.document.getElementById('imgModal').style.display = 'none';
        }}
    }}
    </script>
    ''', unsafe_allow_html=True)

    st.caption(f"🎬 {len(saved_files)} frames saved. Scroll sideways to see all. Click to zoom!")